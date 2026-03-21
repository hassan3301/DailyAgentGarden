"""
Excel export tools for Veloce Agent.
Generates monthly payment/tender-type reports as .xlsx files.
"""

import os
import calendar
import tempfile
import traceback
from datetime import datetime, timedelta

import google.auth
from google.auth.transport import requests as auth_requests
from google.adk.tools import ToolContext
from google.cloud import storage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers

from .config import VELOCE_API_BASE
from .veloce_tools import get_location_id, _api_get

GCS_BUCKET = os.getenv("GOOGLE_CLOUD_STAGING_BUCKET", "gs://daily-agent-bucket").replace("gs://", "")
GCS_REPORTS_FOLDER = "reports"
SIGNED_URL_EXPIRY_HOURS = 1


# Fixed row labels in the order the report expects them.
PAYMENT_TYPE_ROWS = [
    "CASH",
    "DEBIT/CREDIT",
    "GIFT CARD",
    "UBER",
    "SKIP",
    "DOORDASH",
    "UEAT",
    "DEBIT/CREDIT MANUAL",
    "INTERAC",
    "VISA",
    "MASTERCARD",
    "PROMO CARD",
    "AMEX",
    "ROUNDING",
    "LOAD GIFT CARD",
    "PAYMENT TRANSFER",
]

# Map variations returned by the Veloce API to the canonical row labels above.
# Keys are UPPER-CASED API names; values are canonical labels.
_NAME_MAP: dict[str, str] = {
    "CASH":                 "CASH",
    "DEBIT":                "DEBIT/CREDIT",
    "CREDIT":               "DEBIT/CREDIT",
    "DEBIT/CREDIT":         "DEBIT/CREDIT",
    "GIFT CARD":            "GIFT CARD",
    "GIFT":                 "GIFT CARD",
    "UBER":                 "UBER",
    "UBER EATS":            "UBER",
    "SKIP":                 "SKIP",
    "SKIP THE DISHES":      "SKIP",
    "SKIPTHEDISHES":        "SKIP",
    "DOORDASH":             "DOORDASH",
    "DOOR DASH":            "DOORDASH",
    "UEAT":                 "UEAT",
    "U-EAT":                "UEAT",
    "DEBIT/CREDIT MANUAL":  "DEBIT/CREDIT MANUAL",
    "MANUAL":               "DEBIT/CREDIT MANUAL",
    "MANUAL CREDIT":        "DEBIT/CREDIT MANUAL",
    "MANUAL DEBIT":         "DEBIT/CREDIT MANUAL",
    "INTERAC":              "INTERAC",
    "VISA":                 "VISA",
    "MASTERCARD":           "MASTERCARD",
    "MC":                   "MASTERCARD",
    "PROMO CARD":           "PROMO CARD",
    "PROMO":                "PROMO CARD",
    "PROMO CARD PAYMENT":   "PROMO CARD",
    "AMEX":                 "AMEX",
    "AMERICAN EXPRESS":     "AMEX",
    "ROUNDING":             "ROUNDING",
    "LOAD GIFT CARD":       "LOAD GIFT CARD",
    "PAYMENT TRANSFER":     "PAYMENT TRANSFER",
}


def _map_tender_name(api_name: str) -> str:
    """Return the canonical row label for an API tender-type name."""
    import re
    # Normalize: upper-case, collapse whitespace, remove spaces around slashes
    # e.g. "DEBIT / CREDIT MANUAL" -> "DEBIT/CREDIT MANUAL"
    normalized = re.sub(r'\s*/\s*', '/', api_name.strip().upper())
    normalized = re.sub(r'\s+', ' ', normalized)
    return _NAME_MAP.get(normalized, normalized)


def _fetch_day_tender_data(tool_context: ToolContext, location_id: str, date_str: str) -> dict:
    """
    Fetch tender-type breakdown and sales summary for a single day.
    Returns {"tenders": {label: amount, ...}, "gross": ..., "discount": ..., "hst": ..., "tip": ...}
    """
    from_dt = f"{date_str}T00:00:00Z"
    to_dt = f"{date_str}T23:59:59Z"

    # --- tender types ---
    tenders: dict[str, float] = {label: 0.0 for label in PAYMENT_TYPE_ROWS}
    try:
        resp = _api_get(tool_context, f"{VELOCE_API_BASE}/sales/tenderTypes", {
            "locationIDs": [location_id],
            "from": from_dt,
            "to": to_dt,
        })
        resp.raise_for_status()
        for entry in resp.json():
            name = entry.get("nameMain") or entry.get("nameAlt") or ""
            canonical = _map_tender_name(name)
            amount = entry.get("salesAmount", 0)
            if canonical in tenders:
                tenders[canonical] += amount
            else:
                # Unknown type – still capture it
                tenders[canonical] = tenders.get(canonical, 0) + amount
    except Exception as e:
        print(f"WARNING: tender fetch failed for {date_str}: {e}")

    # --- gross & discount from /sales/locations (already aggregated) ---
    gross = 0.0
    discount = 0.0
    try:
        resp = _api_get(tool_context, f"{VELOCE_API_BASE}/sales/locations", {
            "locationIDs": [location_id],
            "from": from_dt,
            "to": to_dt,
            "consolidated": True,
        })
        resp.raise_for_status()
        data = resp.json()
        if data:
            loc = data[0]
            gross = loc.get("salesAmount", 0) or 0
            discount = loc.get("discountsAmount", 0) or 0
    except Exception as e:
        print(f"WARNING: sales/locations fetch failed for {date_str}: {e}")

    # --- hst & tip from invoices (paginated, since days can exceed 100) ---
    hst = 0.0
    tip = 0.0
    try:
        offset = 0
        while True:
            resp = _api_get(tool_context, f"{VELOCE_API_BASE}/invoices", {
                "locationIDs": location_id,
                "from": from_dt,
                "to": to_dt,
                "limit": 100,
                "offset": offset,
            })
            resp.raise_for_status()
            batch = resp.json()
            if not batch:
                break
            for inv in batch:
                hst += inv.get("taxesTotalAmount", 0) or 0
                tip += inv.get("tipsTotalAmount", 0) or 0
            if len(batch) < 100:
                break
            offset += 100
    except Exception as e:
        print(f"WARNING: invoice fetch failed for {date_str}: {e}")

    return {
        "tenders": tenders,
        "gross": gross,
        "discount": discount,
        "hst": hst,
        "tip": tip,
    }


def generate_monthly_payment_report(
    tool_context: ToolContext,
    year: int,
    month: int,
) -> dict:
    """
    Generate a monthly Mode-of-Receipt Excel report.

    Creates an .xlsx file with payment types as rows and each day of the
    month as a column, plus a monthly TOTAL column. Summary rows at the
    bottom show Gross Sale, Discount, HST, Tip, Total Collection, and Diff.

    Args:
        tool_context: ADK context with auth token and location ID
        year: Report year (e.g. 2025)
        month: Report month (1-12)

    Returns:
        Dictionary with status and the file path of the generated Excel file.
    """
    print(f"--- generate_monthly_payment_report: {year}-{month:02d} ---")

    location_id = get_location_id(tool_context)
    location_name = tool_context.state.get("location_name", "Unknown")
    num_days = calendar.monthrange(year, month)[1]

    # ---- Fetch data for every day ------------------------------------------------
    daily_data: list[dict] = []
    for day in range(1, num_days + 1):
        date_str = f"{year}-{month:02d}-{day:02d}"
        print(f"  Fetching {date_str} ...")
        daily_data.append(_fetch_day_tender_data(tool_context, location_id, date_str))

    # ---- Build workbook ----------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    month_name = calendar.month_name[month]
    ws.title = f"{month_name} {year}"

    # Styles
    bold = Font(bold=True)
    currency_fmt = '#,##0.00'
    header_fill_font = Font(bold=True, size=11)
    thin_border = Border(
        bottom=Side(style="thin"),
        top=Side(style="thin"),
        left=Side(style="thin"),
        right=Side(style="thin"),
    )

    # --- Title row ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 2)
    title_cell = ws.cell(row=1, column=1,
                         value=f"{location_name} - Mode of Receipt Report - {month_name} {year}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # --- Header row (row 3): "Mode of Receipt" | day1 | day2 | ... | TOTAL ---
    header_row = 3
    ws.cell(row=header_row, column=1, value="Mode of Receipt").font = header_fill_font
    ws.column_dimensions["A"].width = 22
    for day in range(1, num_days + 1):
        col = day + 1
        cell = ws.cell(row=header_row, column=col,
                       value=f"{month_name[:3]} {day}")
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 12
    total_col = num_days + 2
    ws.cell(row=header_row, column=total_col, value="TOTAL").font = header_fill_font
    ws.column_dimensions[ws.cell(row=header_row, column=total_col).column_letter].width = 14

    # --- Payment type rows ---
    first_data_row = header_row + 1
    for i, label in enumerate(PAYMENT_TYPE_ROWS):
        row = first_data_row + i
        ws.cell(row=row, column=1, value=label).font = bold
        for day_idx, dd in enumerate(daily_data):
            col = day_idx + 2
            val = dd["tenders"].get(label, 0)
            cell = ws.cell(row=row, column=col, value=val)
            cell.number_format = currency_fmt
        # TOTAL column (SUM formula)
        from openpyxl.utils import get_column_letter
        first_col_letter = get_column_letter(2)
        last_col_letter = get_column_letter(num_days + 1)
        total_cell = ws.cell(row=row, column=total_col)
        total_cell.value = f"=SUM({first_col_letter}{row}:{last_col_letter}{row})"
        total_cell.number_format = currency_fmt
        total_cell.font = bold

    # --- Separator row ---
    sep_row = first_data_row + len(PAYMENT_TYPE_ROWS)
    for c in range(1, total_col + 1):
        ws.cell(row=sep_row, column=c).border = Border(bottom=Side(style="medium"))

    # --- Summary rows ---
    summary_labels = ["Gross Sale", "Discount Allowed", "HST", "Tip", "Total Collection", "Diff"]
    summary_keys = ["gross", "discount", "hst", "tip", None, None]
    summary_start = sep_row + 1

    for s_idx, s_label in enumerate(summary_labels):
        row = summary_start + s_idx
        ws.cell(row=row, column=1, value=s_label).font = bold

        if s_label == "Total Collection":
            # Total Collection = Gross Sale - Discount + HST
            gross_row = summary_start        # Gross Sale row
            disc_row = summary_start + 1     # Discount Allowed row
            hst_row = summary_start + 2      # HST row
            for day_idx in range(num_days):
                col = day_idx + 2
                cl = get_column_letter(col)
                ws.cell(row=row, column=col,
                        value=f"={cl}{gross_row}-{cl}{disc_row}+{cl}{hst_row}")
                ws.cell(row=row, column=col).number_format = currency_fmt
        elif s_label == "Diff":
            # Diff = Total Collection - sum of all modes of receipt
            # Should be 0 if payment types account for everything.
            tc_row = summary_start + 4      # Total Collection row
            for day_idx in range(num_days):
                col = day_idx + 2
                cl = get_column_letter(col)
                ws.cell(row=row, column=col,
                        value=f"={cl}{tc_row}-SUM({cl}{first_data_row}:{cl}{first_data_row + len(PAYMENT_TYPE_ROWS) - 1})")
                ws.cell(row=row, column=col).number_format = currency_fmt
        else:
            key = summary_keys[s_idx]
            for day_idx, dd in enumerate(daily_data):
                col = day_idx + 2
                cell = ws.cell(row=row, column=col, value=dd.get(key, 0))
                cell.number_format = currency_fmt

        # TOTAL column for summary rows
        first_col_letter = get_column_letter(2)
        last_col_letter = get_column_letter(num_days + 1)
        total_cell = ws.cell(row=row, column=total_col)
        total_cell.value = f"=SUM({first_col_letter}{row}:{last_col_letter}{row})"
        total_cell.number_format = currency_fmt
        total_cell.font = bold

    # --- Apply thin borders to data area ---
    for r in range(header_row, summary_start + len(summary_labels)):
        for c in range(1, total_col + 1):
            ws.cell(row=r, column=c).border = thin_border

    # ---- Save locally then upload to GCS ----------------------------------------
    out_dir = os.path.join(tempfile.gettempdir(), "veloce_reports")
    os.makedirs(out_dir, exist_ok=True)
    safe_name = location_name.replace(" ", "_").replace("&", "and")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Mode_of_Receipt_{safe_name}_{year}_{month:02d}_{timestamp}.xlsx"
    filepath = os.path.join(out_dir, filename)
    wb.save(filepath)
    print(f"Report saved locally to {filepath}")

    # Upload to GCS and generate signed download URL
    try:
        credentials, project = google.auth.default()
        credentials.refresh(auth_requests.Request())

        gcs_client = storage.Client(credentials=credentials, project=project)
        bucket = gcs_client.bucket(GCS_BUCKET)
        blob_path = f"{GCS_REPORTS_FOLDER}/{filename}"
        blob = bucket.blob(blob_path)
        blob.upload_from_filename(filepath, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        print(f"Uploaded to gs://{GCS_BUCKET}/{blob_path}")

        sign_kwargs = dict(
            version="v4",
            expiration=timedelta(hours=SIGNED_URL_EXPIRY_HOURS),
            method="GET",
            response_disposition=f'attachment; filename="{filename}"',
        )
        # Determine the service account email for IAM-based URL signing.
        # On Vertex AI / Cloud Run it comes from the attached SA credentials.
        # Locally with ADC (user credentials) it must be provided via env var
        # so the IAM signBlob API can be called on behalf of that SA.
        sa_email = getattr(credentials, "service_account_email", None) or os.getenv("GCS_SIGNING_SERVICE_ACCOUNT")
        if sa_email:
            sign_kwargs["service_account_email"] = sa_email
            sign_kwargs["access_token"] = credentials.token
        download_url = blob.generate_signed_url(**sign_kwargs)
        print(f"Signed URL generated (expires in {SIGNED_URL_EXPIRY_HOURS}h)")

        return {
            "status": "success",
            "download_url": download_url,
            "file_path": filepath,
            "gcs_path": f"gs://{GCS_BUCKET}/{blob_path}",
            "message": f"Monthly payment report for {month_name} {year} is ready. Download link (expires in {SIGNED_URL_EXPIRY_HOURS} hour): {download_url}",
            "days_covered": num_days,
            "location": location_name,
        }
    except Exception as e:
        print(f"WARNING: GCS upload failed: {e}")
        traceback.print_exc()
        return {
            "status": "success",
            "file_path": filepath,
            "message": f"Monthly payment report for {month_name} {year} saved to {filepath} (cloud upload failed: {e})",
            "days_covered": num_days,
            "location": location_name,
        }
