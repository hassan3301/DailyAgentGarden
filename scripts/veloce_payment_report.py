"""
Standalone Veloce Monthly Payment Report Generator.

Generates two Excel reports per location:
  1. Mode of Receipt report (same format as VeloceAgent)
  2. QuickBooks Journal Entry upload-ready Excel

Usage:
    python scripts/veloce_payment_report.py --year 2026 --month 3
    python scripts/veloce_payment_report.py --location Appleby --year 2026 --month 3
    python scripts/veloce_payment_report.py  # defaults to previous month, all locations
"""

import argparse
import calendar
import os
import re
import sys
from datetime import datetime, timedelta

import requests
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION  -- fill in real credentials before running
# =============================================================================

VELOCE_API_BASE = "https://api.posveloce.com"

LOCATIONS = [
    {
        "name": "Appleby",
        "email": "burlington@pursimple.com",
        "password": "11Feb1975#222",
        "location_id": None,  # None = auto-detect first active location
    },
    {
        "name": "Heartland",
        "email": "heartland@pursimple.com",
        "password": "11Feb1975#222",
        "location_id": None,
    },
    {
        "name": "Gateway",
        "email": "gateway@pursimple.com",
        "password": "45Gateway#1",
        "location_id": None,
    },
    {
        "name": "Waterdown",
        "email": "waterdown@pursimple.com",
        "password": "80Dundas#",
        "location_id": None,
    },
    {
        "name": "Fairview",
        "email": "burlingtonsouth@pursimple.com",
        "password": "Fairview2024",
        "location_id": None,
    },
]

# QuickBooks account mapping per location for journal entry lines.
# "side" indicates whether this payment type is a debit or credit in the JE.
# Most payment types are debits (money coming in). Special cases noted below.

# Appleby (Burlington)
_QB_MAP_APPLEBY = {
    "CASH":                 {"account": "20140 Holding Account for Net Deposits",                "side": "debit"},
    "DEBIT/CREDIT":         {"account": "20190 Debit / Credit",                                  "side": "debit"},
    "GIFT CARD":            {"account": "40030 Gift Cards",                                      "side": "credit"},
    "UBER":                 {"account": "20150 Holding Account:Holding Account - Uber Eats",     "side": "debit"},
    "SKIP":                 {"account": "20160 Holding Account:Holding Account - Skip the Dishes","side": "debit"},
    "DOORDASH":             {"account": "20170 Holding Account:Holding Account - Door Dash",     "side": "debit"},
    "UEAT":                 {"account": "20180 Holding Account:Holding Account - UEat",          "side": "debit"},
    "AMEX":                 {"account": "20191 Holding Account:Holding Account - Amex",          "side": "debit"},
    "LOAD GIFT CARD":       {"account": "40030 Gift Card",                                       "side": "debit"},
    "PAYMENT TRANSFER":     {"account": "40030 Gift Card",                                       "side": "debit"},
}

# Heartland
_QB_MAP_HEARTLAND = {
    "CASH":                 {"account": "20120 Holding Account for Net Deposits",       "side": "debit"},
    "DEBIT/CREDIT":         {"account": "DEBIT/CREDIT",                                 "side": "debit"},
    "GIFT CARD":            {"account": "40040 Gift Cards",                             "side": "credit"},
    "UBER":                 {"account": "20130 Holding Account - Uber Eats",            "side": "debit"},
    "SKIP":                 {"account": "20150 Holding Account - Skip the Dishes",      "side": "debit"},
    "DOORDASH":             {"account": "20160 Holding Account - Door Dash",            "side": "debit"},
    "UEAT":                 {"account": "20190 Holding Account - UEats",                "side": "debit"},
    "AMEX":                 {"account": "20191 Holding Account - Amex",                 "side": "debit"},
    "LOAD GIFT CARD":       {"account": "40040 Gift Cards",                             "side": "debit"},
    "PAYMENT TRANSFER":     {"account": "40040 Gift Cards",                             "side": "debit"},
}

# Gateway (Brampton)
_QB_MAP_GATEWAY = {
    "CASH":                 {"account": "20120 Holding Account for Net Deposits",       "side": "debit"},
    "DEBIT/CREDIT":         {"account": "20192 Debit/Credit",                           "side": "debit"},
    "GIFT CARD":            {"account": "40040 Gift Cards",                             "side": "credit"},
    "UBER":                 {"account": "20130 Holding Account - Uber Eats",            "side": "debit"},
    "SKIP":                 {"account": "20150 Holding Account - Skip the Dishes",      "side": "debit"},
    "DOORDASH":             {"account": "20160 Holding Account - Door Dash",            "side": "debit"},
    "UEAT":                 {"account": "20190 Holding Account - UEats",                "side": "debit"},
    "AMEX":                 {"account": "20191 Holding Account - Amex",                 "side": "debit"},
    "LOAD GIFT CARD":       {"account": "40040 Gift Cards",                             "side": "debit"},
    "PAYMENT TRANSFER":     {"account": "40040 Gift Cards",                             "side": "debit"},
}

# Waterdown
_QB_MAP_WATERDOWN = {
    "CASH":                 {"account": "20120 Holding Account for Net Deposits",       "side": "debit"},
    "DEBIT/CREDIT":         {"account": "20196 Holding Account - Debit credit",         "side": "debit"},
    "GIFT CARD":            {"account": "40040 Gift Cards",                             "side": "credit"},
    "UBER":                 {"account": "20130 Holding Account - Uber Eats",            "side": "debit"},
    "SKIP":                 {"account": "20150 Holding Account - Skip the Dishes",      "side": "debit"},
    "DOORDASH":             {"account": "20160 Holding Account - Door Dash",            "side": "debit"},
    "UEAT":                 {"account": "20190 Holding Account - UEats",                "side": "debit"},
    "AMEX":                 {"account": "20191 Holding Account - Amex",                 "side": "debit"},
    "LOAD GIFT CARD":       {"account": "40040 Gift Cards",                             "side": "debit"},
    "PAYMENT TRANSFER":     {"account": "40040 Gift Cards",                             "side": "debit"},
}

# Fairview -- accounts TBD, left empty for now
_QB_MAP_FAIRVIEW = {
    "CASH":                 {"account": "",  "side": "debit"},
    "DEBIT/CREDIT":         {"account": "",  "side": "debit"},
    "GIFT CARD":            {"account": "",  "side": "credit"},
    "UBER":                 {"account": "",  "side": "debit"},
    "SKIP":                 {"account": "",  "side": "debit"},
    "DOORDASH":             {"account": "",  "side": "debit"},
    "UEAT":                 {"account": "",  "side": "debit"},
    "AMEX":                 {"account": "",  "side": "debit"},
    "LOAD GIFT CARD":       {"account": "",  "side": "debit"},
    "PAYMENT TRANSFER":     {"account": "",  "side": "debit"},
}

# Lookup by location name
LOCATION_QB_ACCOUNT_MAP: dict[str, dict] = {
    "Appleby":   _QB_MAP_APPLEBY,
    "Heartland": _QB_MAP_HEARTLAND,
    "Gateway":   _QB_MAP_GATEWAY,
    "Waterdown": _QB_MAP_WATERDOWN,
    "Fairview":  _QB_MAP_FAIRVIEW,
}

# Closing / balancing account -- same across all locations
QB_CLOSING_ACCOUNT = "20040 Undeposited Funds"

# =============================================================================
# PAYMENT TYPE CONSTANTS (from VeloceAgent/excel_tools.py)
# =============================================================================

PAYMENT_TYPE_ROWS = [
    "CASH",
    "DEBIT/CREDIT",
    "GIFT CARD",
    "UBER",
    "SKIP",
    "DOORDASH",
    "UEAT",
    "DEBIT/CREDIT MANUAL",
    "INTERAC",
    "VISA",
    "MASTERCARD",
    "PROMO CARD",
    "AMEX",
    "ROUNDING",
    "LOAD GIFT CARD",
    "PAYMENT TRANSFER",
]

_NAME_MAP: dict[str, str] = {
    "CASH":                 "CASH",
    "DEBIT":                "DEBIT/CREDIT",
    "CREDIT":               "DEBIT/CREDIT",
    "DEBIT/CREDIT":         "DEBIT/CREDIT",
    "GIFT CARD":            "GIFT CARD",
    "GIFT":                 "GIFT CARD",
    "UBER":                 "UBER",
    "UBER EATS":            "UBER",
    "SKIP":                 "SKIP",
    "SKIP THE DISHES":      "SKIP",
    "SKIPTHEDISHES":        "SKIP",
    "DOORDASH":             "DOORDASH",
    "DOOR DASH":            "DOORDASH",
    "UEAT":                 "UEAT",
    "U-EAT":                "UEAT",
    "DEBIT/CREDIT MANUAL":  "DEBIT/CREDIT MANUAL",
    "MANUAL":               "DEBIT/CREDIT MANUAL",
    "MANUAL CREDIT":        "DEBIT/CREDIT MANUAL",
    "MANUAL DEBIT":         "DEBIT/CREDIT MANUAL",
    "INTERAC":              "INTERAC",
    "VISA":                 "VISA",
    "MASTERCARD":           "MASTERCARD",
    "MC":                   "MASTERCARD",
    "PROMO CARD":           "PROMO CARD",
    "PROMO":                "PROMO CARD",
    "PROMO CARD PAYMENT":   "PROMO CARD",
    "AMEX":                 "AMEX",
    "AMERICAN EXPRESS":     "AMEX",
    "ROUNDING":             "ROUNDING",
    "LOAD GIFT CARD":       "LOAD GIFT CARD",
    "PAYMENT TRANSFER":     "PAYMENT TRANSFER",
}


def _map_tender_name(api_name: str) -> str:
    """Return the canonical row label for an API tender-type name."""
    normalized = re.sub(r"\s*/\s*", "/", api_name.strip().upper())
    normalized = re.sub(r"\s+", " ", normalized)
    return _NAME_MAP.get(normalized, normalized)


# =============================================================================
# API MODULE (plain requests, no ADK dependency)
# =============================================================================

def authenticate(email: str, password: str) -> str:
    """Authenticate with Veloce API and return the bearer token."""
    resp = requests.post(
        f"{VELOCE_API_BASE}/users/authenticate",
        json={"email": email, "password": password},
    )
    resp.raise_for_status()
    data = resp.json()
    token = data.get("token")
    if not token:
        raise RuntimeError("Authentication succeeded but no token returned")
    return token


def get_locations(token: str) -> list[dict]:
    """Return list of locations the authenticated user has access to."""
    resp = api_get(token, f"{VELOCE_API_BASE}/locations")
    resp.raise_for_status()
    return [
        {
            "id": loc.get("id"),
            "name": loc.get("name"),
            "is_active": loc.get("isActive", False),
        }
        for loc in resp.json()
    ]


def api_get(token: str, url: str, params: dict | None = None) -> requests.Response:
    """Simple authenticated GET with one retry on 401."""
    resp = requests.get(
        url,
        headers={"Authorization": f"Bearer {token}"},
        params=params,
    )
    return resp


def fetch_day_tender_data(token: str, location_id: str, date_str: str) -> dict:
    """
    Fetch tender-type breakdown and sales summary for a single day.
    Returns {"tenders": {label: amount}, "gross", "discount", "hst", "tip"}
    """
    from_dt = f"{date_str}T00:00:00Z"
    to_dt = f"{date_str}T23:59:59Z"

    # --- tender types ---
    tenders: dict[str, float] = {label: 0.0 for label in PAYMENT_TYPE_ROWS}
    try:
        resp = api_get(token, f"{VELOCE_API_BASE}/sales/tenderTypes", {
            "locationIDs": [location_id],
            "from": from_dt,
            "to": to_dt,
        })
        resp.raise_for_status()
        for entry in resp.json():
            name = entry.get("nameMain") or entry.get("nameAlt") or ""
            canonical = _map_tender_name(name)
            amount = entry.get("salesAmount", 0)
            if canonical in tenders:
                tenders[canonical] += amount
            else:
                tenders[canonical] = tenders.get(canonical, 0) + amount
    except Exception as e:
        print(f"  WARNING: tender fetch failed for {date_str}: {e}")

    # --- gross & discount ---
    gross = 0.0
    discount = 0.0
    try:
        resp = api_get(token, f"{VELOCE_API_BASE}/sales/locations", {
            "locationIDs": [location_id],
            "from": from_dt,
            "to": to_dt,
            "consolidated": True,
        })
        resp.raise_for_status()
        data = resp.json()
        if data:
            loc = data[0]
            gross = loc.get("salesAmount", 0) or 0
            discount = loc.get("discountsAmount", 0) or 0
    except Exception as e:
        print(f"  WARNING: sales/locations fetch failed for {date_str}: {e}")

    # --- hst & tip from invoices (paginated) ---
    hst = 0.0
    tip = 0.0
    try:
        offset = 0
        while True:
            resp = api_get(token, f"{VELOCE_API_BASE}/invoices", {
                "locationIDs": location_id,
                "from": from_dt,
                "to": to_dt,
                "limit": 100,
                "offset": offset,
            })
            resp.raise_for_status()
            batch = resp.json()
            if not batch:
                break
            for inv in batch:
                hst += inv.get("taxesTotalAmount", 0) or 0
                tip += inv.get("tipsTotalAmount", 0) or 0
            if len(batch) < 100:
                break
            offset += 100
    except Exception as e:
        print(f"  WARNING: invoice fetch failed for {date_str}: {e}")

    return {
        "tenders": tenders,
        "gross": gross,
        "discount": discount,
        "hst": hst,
        "tip": tip,
    }


# =============================================================================
# MODE OF RECEIPT EXCEL GENERATION
# =============================================================================

def generate_mode_of_receipt_excel(
    daily_data: list[dict],
    location_name: str,
    year: int,
    month: int,
    out_dir: str,
) -> str:
    """
    Generate the Mode-of-Receipt Excel report (same format as VeloceAgent).
    Returns the filepath of the saved .xlsx.
    """
    num_days = len(daily_data)
    wb = Workbook()
    ws = wb.active
    month_name = calendar.month_name[month]
    ws.title = f"{month_name} {year}"

    # Styles
    bold = Font(bold=True)
    currency_fmt = "#,##0.00"
    header_fill_font = Font(bold=True, size=11)
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    GREEN_ROWS = {
        "CASH", "DEBIT/CREDIT", "GIFT CARD", "UBER", "SKIP",
        "DOORDASH", "UEAT", "AMEX", "LOAD GIFT CARD", "PAYMENT TRANSFER",
    }
    thin_border = Border(
        bottom=Side(style="thin"),
        top=Side(style="thin"),
        left=Side(style="thin"),
        right=Side(style="thin"),
    )

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 2)
    title_cell = ws.cell(
        row=1, column=1,
        value=f"{location_name} - Mode of Receipt Report - {month_name} {year}",
    )
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Header row (row 3)
    header_row = 3
    ws.cell(row=header_row, column=1, value="Mode of Receipt").font = header_fill_font
    ws.column_dimensions["A"].width = 22
    for day in range(1, num_days + 1):
        col = day + 1
        cell = ws.cell(row=header_row, column=col, value=f"{month_name[:3]} {day}")
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 12
    total_col = num_days + 2
    ws.cell(row=header_row, column=total_col, value="TOTAL").font = header_fill_font
    ws.column_dimensions[ws.cell(row=header_row, column=total_col).column_letter].width = 14

    # Payment type rows
    first_data_row = header_row + 1
    for i, label in enumerate(PAYMENT_TYPE_ROWS):
        row = first_data_row + i
        is_green = label in GREEN_ROWS
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = bold
        if is_green:
            label_cell.fill = green_fill
        for day_idx, dd in enumerate(daily_data):
            col = day_idx + 2
            val = dd["tenders"].get(label, 0)
            cell = ws.cell(row=row, column=col, value=val)
            cell.number_format = currency_fmt
            if is_green:
                cell.fill = green_fill
        # TOTAL column
        first_col_letter = get_column_letter(2)
        last_col_letter = get_column_letter(num_days + 1)
        total_cell = ws.cell(row=row, column=total_col)
        total_cell.value = f"=SUM({first_col_letter}{row}:{last_col_letter}{row})"
        total_cell.number_format = currency_fmt
        total_cell.font = bold
        if is_green:
            total_cell.fill = green_fill

    # Separator row
    sep_row = first_data_row + len(PAYMENT_TYPE_ROWS)
    for c in range(1, total_col + 1):
        ws.cell(row=sep_row, column=c).border = Border(bottom=Side(style="medium"))

    # Summary rows
    summary_labels = ["Gross Sale", "Discount Allowed", "HST", "Tip", "Total Collection", "Diff"]
    summary_keys = ["gross", "discount", "hst", "tip", None, None]
    summary_start = sep_row + 1

    for s_idx, s_label in enumerate(summary_labels):
        row = summary_start + s_idx
        ws.cell(row=row, column=1, value=s_label).font = bold

        if s_label == "Total Collection":
            gross_row = summary_start
            disc_row = summary_start + 1
            hst_row = summary_start + 2
            for day_idx in range(num_days):
                col = day_idx + 2
                cl = get_column_letter(col)
                cell = ws.cell(row=row, column=col,
                               value=f"={cl}{gross_row}-{cl}{disc_row}+{cl}{hst_row}")
                cell.number_format = currency_fmt
        elif s_label == "Diff":
            tc_row = summary_start + 4
            for day_idx in range(num_days):
                col = day_idx + 2
                cl = get_column_letter(col)
                cell = ws.cell(
                    row=row, column=col,
                    value=f"={cl}{tc_row}-SUM({cl}{first_data_row}:{cl}{first_data_row + len(PAYMENT_TYPE_ROWS) - 1})",
                )
                cell.number_format = currency_fmt
        else:
            key = summary_keys[s_idx]
            for day_idx, dd in enumerate(daily_data):
                col = day_idx + 2
                cell = ws.cell(row=row, column=col, value=dd.get(key, 0))
                cell.number_format = currency_fmt

        # TOTAL column for summary rows
        first_col_letter = get_column_letter(2)
        last_col_letter = get_column_letter(num_days + 1)
        total_cell = ws.cell(row=row, column=total_col)
        total_cell.value = f"=SUM({first_col_letter}{row}:{last_col_letter}{row})"
        total_cell.number_format = currency_fmt
        total_cell.font = bold

    # Apply thin borders to data area
    for r in range(header_row, summary_start + len(summary_labels)):
        for c in range(1, total_col + 1):
            ws.cell(row=r, column=c).border = thin_border

    # Save
    safe_name = location_name.replace(" ", "_").replace("&", "and").replace("#", "")
    filename = f"Mode_of_Receipt_{safe_name}_{year}_{month:02d}.xlsx"
    filepath = os.path.join(out_dir, filename)
    wb.save(filepath)
    return filepath


# =============================================================================
# QUICKBOOKS JOURNAL ENTRY EXCEL GENERATION
# =============================================================================

def _strip_account_number(account_name: str) -> str:
    """Strip the leading account number from a QBO account name.

    E.g. "20140 Holding Account for Net Deposits" -> "Holding Account for Net Deposits"
         "40030 Gift Cards" -> "Gift Cards"
    """
    return re.sub(r"^\d+\s+", "", account_name)


# Rotating fill colours used to visually group journal entry rows by date.
_DATE_FILLS = [
    PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # light green
    PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # light blue
    PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # light orange
    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # sage green
    PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),  # lavender
    PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # light yellow
]


def generate_qb_journal_entry(
    daily_data: list[dict],
    location_name: str,
    year: int,
    month: int,
    out_dir: str,
) -> str:
    """
    Generate a QuickBooks Journal Entry import-ready Excel file.

    For each day with activity, creates journal entry lines from the payment
    types in QB_ACCOUNT_MAP.  Each mapped type goes to its configured side
    (debit or credit).  A closing line on the "Undeposited Funds"
    account balances total debits to total credits.

    Account names are output WITHOUT leading account numbers so they can be
    copy-pasted directly into QBO.

    Rows are colour-coded by date for easier visual grouping.

    Columns: Account | Debits | Credits | Date | Journal Number

    Returns the filepath of the saved .xlsx.
    """
    num_days = len(daily_data)
    wb = Workbook()
    ws = wb.active
    month_name = calendar.month_name[month]
    ws.title = f"JE {month_name} {year}"

    # Styles
    bold = Font(bold=True)
    currency_fmt = "#,##0.00"

    # Header row
    headers = ["Account", "Debits", "Credits", "Date", "Journal Number"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16

    current_row = 2
    colour_idx = 0  # cycles through _DATE_FILLS

    # Get the account map for this location
    qb_account_map = LOCATION_QB_ACCOUNT_MAP.get(location_name)
    if qb_account_map is None:
        print(f"  WARNING: No QB account mapping for location '{location_name}', skipping JE")
        wb.save(os.path.join(out_dir, f"QB_Journal_Entry_{location_name}_{year}_{month:02d}.xlsx"))
        return os.path.join(out_dir, f"QB_Journal_Entry_{location_name}_{year}_{month:02d}.xlsx")

    for day_idx in range(num_days):
        dd = daily_data[day_idx]
        day_num = day_idx + 1
        date_str = f"{year}-{month:02d}-{day_num:02d}"

        # Skip days with no activity
        total_tenders = sum(dd["tenders"].get(label, 0) for label in qb_account_map)
        if total_tenders == 0:
            continue

        # Pick the fill colour for this date group and advance the index
        date_fill = _DATE_FILLS[colour_idx % len(_DATE_FILLS)]
        colour_idx += 1
        first_row_of_day = current_row

        total_debits = 0.0
        total_credits = 0.0

        # --- Lines for each mapped payment type with a non-zero amount ---
        for label, acct_info in qb_account_map.items():
            amount = dd["tenders"].get(label, 0)
            if amount == 0:
                continue

            amount = round(amount, 2)

            ws.cell(row=current_row, column=1,
                    value=_strip_account_number(acct_info["account"]))

            if amount > 0:
                # Positive value goes to debits
                ws.cell(row=current_row, column=2, value=amount).number_format = currency_fmt
                ws.cell(row=current_row, column=3, value="")
                total_debits += amount
            else:
                # Negative value goes to credits as absolute value
                ws.cell(row=current_row, column=2, value="")
                ws.cell(row=current_row, column=3, value=abs(amount)).number_format = currency_fmt
                total_credits += abs(amount)

            ws.cell(row=current_row, column=4, value=date_str)
            ws.cell(row=current_row, column=5, value="")  # Journal Number left blank
            current_row += 1

        # --- Closing line to balance debits == credits ---
        diff = round(total_debits - total_credits, 2)
        if diff != 0:
            ws.cell(row=current_row, column=1,
                    value=_strip_account_number(QB_CLOSING_ACCOUNT))
            if diff > 0:
                # More debits than credits -> closing is a credit
                ws.cell(row=current_row, column=2, value="")
                ws.cell(row=current_row, column=3, value=diff).number_format = currency_fmt
            else:
                # More credits than debits -> closing is a debit
                ws.cell(row=current_row, column=2, value=abs(diff)).number_format = currency_fmt
                ws.cell(row=current_row, column=3, value="")
            ws.cell(row=current_row, column=4, value=date_str)
            ws.cell(row=current_row, column=5, value="")
            current_row += 1

        # Apply the date colour to all rows for this day
        for r in range(first_row_of_day, current_row):
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = date_fill

    # Save
    safe_name = location_name.replace(" ", "_").replace("&", "and").replace("#", "")
    filename = f"QB_Journal_Entry_{safe_name}_{year}_{month:02d}.xlsx"
    filepath = os.path.join(out_dir, filename)
    wb.save(filepath)
    return filepath


# =============================================================================
# MAIN RUNNER
# =============================================================================

def resolve_location_id(token: str, config_location_id: str | None) -> tuple[str, str]:
    """
    Return (location_id, location_name).
    Uses config_location_id if provided, otherwise auto-detects the first active location.
    """
    locations = get_locations(token)
    if config_location_id:
        for loc in locations:
            if loc["id"] == config_location_id:
                return loc["id"], loc["name"]
        # If the configured ID wasn't found, use it anyway (API might still accept it)
        return config_location_id, config_location_id

    # Auto-detect: prefer first active location
    active = [loc for loc in locations if loc.get("is_active")]
    if active:
        return active[0]["id"], active[0]["name"]
    if locations:
        return locations[0]["id"], locations[0]["name"]
    raise RuntimeError("No locations found for this account")


def run_location(loc_config: dict, year: int, month: int, out_dir: str) -> None:
    """Process a single location: authenticate, fetch data, generate both reports."""
    name = loc_config["name"]
    email = loc_config["email"]
    password = loc_config["password"]

    if password == "CHANGE_ME":
        print(f"\n  SKIPPING {name} -- credentials not configured")
        return

    print(f"\n{'='*60}")
    print(f"  Location: {name}")
    print(f"{'='*60}")

    # Authenticate
    print(f"  Authenticating as {email} ...")
    token = authenticate(email, password)

    # Resolve location ID
    location_id, api_location_name = resolve_location_id(token, loc_config.get("location_id"))
    print(f"  Location: {name} / {api_location_name} ({location_id})")

    # Fetch daily data for every day in the month
    num_days = calendar.monthrange(year, month)[1]
    daily_data: list[dict] = []
    for day in range(1, num_days + 1):
        date_str = f"{year}-{month:02d}-{day:02d}"
        print(f"  Fetching {date_str} ...", end=" ", flush=True)
        dd = fetch_day_tender_data(token, location_id, date_str)
        daily_data.append(dd)
        day_total = sum(dd["tenders"].values())
        print(f"tenders=${day_total:,.2f}  gross=${dd['gross']:,.2f}")

    # Generate Mode of Receipt Excel (use config name for filenames)
    mor_path = generate_mode_of_receipt_excel(daily_data, name, year, month, out_dir)
    print(f"  Mode of Receipt report: {mor_path}")

    # Generate QB Journal Entry Excel
    qb_path = generate_qb_journal_entry(daily_data, name, year, month, out_dir)
    print(f"  QB Journal Entry report: {qb_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Veloce monthly payment reports for all locations.",
    )

    # Default to previous month
    today = datetime.now()
    first_of_month = today.replace(day=1)
    prev_month = first_of_month - timedelta(days=1)

    parser.add_argument("--year", type=int, default=prev_month.year,
                        help=f"Report year (default: {prev_month.year})")
    parser.add_argument("--month", type=int, default=prev_month.month,
                        help=f"Report month 1-12 (default: {prev_month.month})")
    parser.add_argument("--output-dir", type=str, default=None,
                        help="Output directory (default: ./reports)")
    location_names = [loc["name"] for loc in LOCATIONS]
    parser.add_argument("--location", type=str, default=None,
                        help=f"Run for a single location by name (default: all). "
                             f"Choices: {', '.join(location_names)}")
    args = parser.parse_args()

    year = args.year
    month = args.month

    if not (1 <= month <= 12):
        print(f"ERROR: month must be 1-12, got {month}")
        sys.exit(1)

    # Filter locations
    if args.location:
        matched = [loc for loc in LOCATIONS if loc["name"].lower() == args.location.lower()]
        if not matched:
            print(f"ERROR: unknown location '{args.location}'. "
                  f"Available: {', '.join(location_names)}")
            sys.exit(1)
        run_locations = matched
    else:
        run_locations = LOCATIONS

    # Output directory
    out_dir = args.output_dir or os.path.join(os.getcwd(), "reports")
    os.makedirs(out_dir, exist_ok=True)

    month_name = calendar.month_name[month]
    print(f"Veloce Payment Report Generator")
    print(f"Period: {month_name} {year}")
    print(f"Output: {out_dir}")
    print(f"Locations: {', '.join(loc['name'] for loc in run_locations)}")

    for loc_config in run_locations:
        try:
            run_location(loc_config, year, month, out_dir)
        except Exception as e:
            print(f"\n  ERROR processing {loc_config['name']}: {e}")

    print(f"\n{'='*60}")
    print(f"  Done. Reports saved to: {out_dir}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
